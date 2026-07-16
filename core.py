# -*- coding: utf-8 -*-
"""
Tang du lieu: SQLite, doc/chuan hoa Excel, xuat CSV/Excel.
Khong phu thuoc vao thu vien giao dien - dung chung cho moi UI.
"""
import os
import re
import csv
import sys
import json
import shutil
import secrets
import sqlite3
import hashlib
import datetime
import unicodedata
import urllib.request

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

if getattr(sys, "frozen", False):
    # Khi da dong goi bang PyInstaller, __file__ tro vao trong thu muc _internal
    # (bi xoa/thay the moi lan cap nhat) - phai dung thu muc chua file .exe thuc
    # su de benh_nhan.db khong bi mat khi update.
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "benh_nhan.db")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
PASSWORD_FILE = os.path.join(BASE_DIR, "app_password.hash")
VERSION_FILE = os.path.join(BASE_DIR, "VERSION.txt")
UPDATE_TOKEN_FILE = os.path.join(BASE_DIR, "update_token.txt")
GITHUB_OWNER = "Monsterph6"
GITHUB_REPO = "QuanlybenhKLN"

COLUMNS = [
    ("id", "ID"),
    ("tt", "TT"),
    ("ho_ten", "Họ và tên"),
    ("gioi_tinh", "Giới tính"),
    ("nam_sinh_raw", "Ngày sinh"),
    ("birth_year", "Năm sinh"),
    ("ma_bhyt", "Mã BHYT"),
    ("so_cccd", "Số CCCD"),
    ("dia_chi", "Địa chỉ"),
    ("phuong_xa", "Phường/Xã"),
    ("tinh_tp", "Tỉnh/TP"),
    ("ngay_kham_raw", "Ngày khám"),
    ("chan_doan", "Chẩn đoán"),
    ("benh", "Nhóm bệnh (KLN)"),
    ("benh_kem_theo", "Bệnh kèm theo"),
    ("nguon_file", "Nguồn nhập"),
    ("lich_su_kham", "Lịch sử khám (đã gộp)"),
]

# ------------------------------------------------------------------
# Phan loai nhom benh khong lay nhiem (KLN/NCD) tu cot Chan doan (tu do,
# khong co cau truc) - dung khi nhap Excel va cho nut "Xac dinh lai nhom
# benh" o tab Nhap du lieu. Danh sach nay CHUA phai chuan y khoa day du -
# chi la tu khoa doi chieu don gian, nguoi dung can xem lai/sua tay khi
# can do chinh xac cao (vd bao cao len tuyen tren).
#
# Moi phan tu: (ma, nhan hien thi, danh sach tu khoa - da bo dau, viet
# thuong, dung de doi chieu voi Chan doan cung da bo dau/viet thuong).
# Mot benh nhan co the thuoc NHIEU nhom cung luc (vd vua THA vua DTD) -
# cot "benh" luu danh sach nhan cach nhau boi ", ".
# ------------------------------------------------------------------

DISEASE_CATEGORIES = [
    ("THA", "Tăng huyết áp", [
        "tang huyet ap", "cao huyet ap", "huyet ap cao", "tha",
    ]),
    ("DTD", "Đái tháo đường", [
        "dai thao duong", "tieu duong", "dtd",
    ]),
    ("COPD_HEN", "COPD / Hen phế quản", [
        "copd", "phoi tac nghen man tinh", "hen phe quan", "hen suyen",
    ]),
    ("UNGTHU", "Ung thư", [
        "ung thu",
    ]),
    ("TAMTHAN", "Tâm thần", [
        "tam than", "tram cam", "roi loan lo au",
    ]),
]
DISEASE_OTHER_LABEL = "Khác"


def _normalize_text(s):
    """Bo dau tieng Viet + viet thuong, dung de doi chieu tu khoa khong
    phu thuoc file nguon co go dau day du hay khong."""
    if not s:
        return ""
    s = s.replace("đ", "d").replace("Đ", "D")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower()


def infer_diseases(chan_doan_text):
    """Doi chieu tu khoa trong Chan doan voi DISEASE_CATEGORIES, tra ve
    chuoi cac nhan benh khop nhau cach nhau boi ", " (co the nhieu nhom
    cung luc). Neu Chan doan co noi dung nhung khong khop nhom nao, tra ve
    DISEASE_OTHER_LABEL ("Khác") de khong bo sot du lieu. Neu Chan doan
    rong, tra ve chuoi rong.

    Doi chieu theo RANH GIOI TU (\\b...\\b), khong phai chi la substring -
    can thiet vi co tu viet tat ngan ("tha", "dtd") de nhan dang duoc kieu
    ghi tat pho bien trong benh an, nhung neu chi so sanh substring thi se
    khop nham vao cac tu khac chua chuoi do (vd "tha" nam trong "thắt" cua
    "đau thắt ngực")."""
    text = _normalize_text(chan_doan_text)
    if not text.strip():
        return ""
    matched = [
        label for _, label, keywords in DISEASE_CATEGORIES
        if any(re.search(r"\b" + re.escape(kw) + r"\b", text) for kw in keywords)
    ]
    return ", ".join(matched) if matched else DISEASE_OTHER_LABEL

# Cac truong nguoi dung co the chon (ket hop bang AND) de xac dinh "trung nhau".
# Moi phan tu: (nhan hien thi, bieu thuc dung trong khoa gop nhom, dieu kien khac rong)
DEDUP_FIELDS = [
    ("Số CCCD", "TRIM(so_cccd)", "TRIM(so_cccd) <> ''"),
    ("Mã BHYT", "TRIM(ma_bhyt)", "TRIM(ma_bhyt) <> ''"),
    ("Họ và tên", "UPPER(TRIM(ho_ten))", "TRIM(ho_ten) <> ''"),
    ("Năm sinh", "IFNULL(birth_year,'')", "birth_year IS NOT NULL"),
    ("Giới tính", "gioi_tinh", "TRIM(gioi_tinh) <> ''"),
    ("Địa chỉ", "UPPER(TRIM(dia_chi))", "TRIM(dia_chi) <> ''"),
    ("Phường/Xã", "UPPER(TRIM(phuong_xa))", "TRIM(phuong_xa) <> ''"),
    ("Tỉnh/TP", "UPPER(TRIM(tinh_tp))", "TRIM(tinh_tp) <> ''"),
]

PAGE_SIZE = 200


def build_dedup_key(selected_labels):
    """selected_labels: danh sach nhan (theo DEDUP_FIELDS) nguoi dung da chon,
    ket hop bang AND de xac dinh 2 ban ghi la "trung nhau".
    Tra ve (key_expr, key_where, key_type) dung cho cac ham loc trung ben duoi."""
    chosen = [f for f in DEDUP_FIELDS if f[0] in selected_labels]
    if not chosen:
        raise ValueError("Chưa chọn trường nào để xác định trùng.")
    key_expr = " || '#' || ".join(c[1] for c in chosen)
    key_where = " AND ".join(c[2] for c in chosen)
    key_type = "+".join(c[0] for c in chosen)
    return key_expr, key_where, key_type


# ------------------------------------------------------------------
# Cau hinh mang LAN (may chu / may tram) - chia se benh_nhan.db cho
# nhieu may trong cung mang noi bo, khong dung Internet/cloud.
# ------------------------------------------------------------------

LAN_CONFIG_FILE = os.path.join(BASE_DIR, "lan_config.json")

REMOTE_BASE_URL = None
REMOTE_API_KEY = ""


def load_lan_config():
    """Doc cau hinh che do mang LAN da luu, vi du:
    {"role": "server", "port": 8765, "api_key": ""}
    {"role": "client", "server_url": "http://192.168.1.10:8765", "api_key": ""}
    role mac dinh "single" (1 may, khong chia se qua mang)."""
    if os.path.exists(LAN_CONFIG_FILE):
        try:
            with open(LAN_CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (OSError, ValueError):
            pass
    return {"role": "single"}


def save_lan_config(cfg):
    with open(LAN_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def get_lan_api_key():
    return (load_lan_config().get("api_key") or "").strip()


def configure_remote(url, api_key=""):
    """Bat che do may tram: moi thao tac CSDL duoc goi qua mang LAN toi may
    chu tai `url` (vi du 'http://192.168.1.10:8765') thay vi mo file
    benh_nhan.db cuc bo. Goi voi url rong/None de tat (dung file cuc bo)."""
    global REMOTE_BASE_URL, REMOTE_API_KEY
    REMOTE_BASE_URL = url.rstrip("/") if url else None
    REMOTE_API_KEY = api_key or ""


def is_remote():
    return REMOTE_BASE_URL is not None


# ------------------------------------------------------------------
# Kiem soat IP ket noi toi may chu LAN (whitelist) - chi dung o phia may
# chu (netserver.py kiem tra, server_tray.py quan ly qua menu "Quản lý IP
# được phép kết nối"). Khong lien quan gi toi may tram.
# ------------------------------------------------------------------

ACL_CONFIG_FILE = os.path.join(BASE_DIR, "acl_config.json")


def load_acl_config():
    """Doc cau hinh kiem soat IP da luu, vi du:
    {"mode": "allow_all", "allowed_ips": []}
    {"mode": "whitelist", "allowed_ips": ["192.168.1.11", "192.168.1.12"]}
    mode mac dinh "allow_all" (khong gioi han - giu hanh vi cu neu chua
    tung cau hinh)."""
    if os.path.exists(ACL_CONFIG_FILE):
        try:
            with open(ACL_CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            cfg.setdefault("mode", "allow_all")
            cfg.setdefault("allowed_ips", [])
            return cfg
        except (OSError, ValueError):
            pass
    return {"mode": "allow_all", "allowed_ips": []}


def save_acl_config(cfg):
    with open(ACL_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ------------------------------------------------------------------
# Tang du lieu (DB)
# ------------------------------------------------------------------

def get_conn():
    if REMOTE_BASE_URL:
        from netclient import RemoteConnection
        return RemoteConnection(REMOTE_BASE_URL, REMOTE_API_KEY)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS patients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tt INTEGER,
            ho_ten TEXT,
            gioi_tinh TEXT,
            nam_sinh_raw TEXT,
            birth_year INTEGER,
            ma_bhyt TEXT,
            so_cccd TEXT,
            dia_chi TEXT,
            phuong_xa TEXT,
            tinh_tp TEXT,
            ngay_kham_raw TEXT,
            ngay_kham_date TEXT,
            chan_doan TEXT,
            benh_kem_theo TEXT,
            nguon_file TEXT,
            imported_at TEXT,
            row_hash TEXT UNIQUE
        )
    """)
    existing_cols = {r["name"] for r in conn.execute("PRAGMA table_info(patients)")}
    if "lich_su_kham" not in existing_cols:
        conn.execute("ALTER TABLE patients ADD COLUMN lich_su_kham TEXT")
    if "benh" not in existing_cols:
        conn.execute("ALTER TABLE patients ADD COLUMN benh TEXT")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cccd ON patients(so_cccd)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_bhyt ON patients(ma_bhyt)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_hoten ON patients(ho_ten COLLATE NOCASE)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_birthyear ON patients(birth_year)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS dedup_exceptions (
            key_type TEXT NOT NULL,
            key_value TEXT NOT NULL,
            ten_dai_dien TEXT,
            created_at TEXT,
            PRIMARY KEY (key_type, key_value)
        )
    """)
    conn.commit()
    conn.close()


# ------------------------------------------------------------------
# Ngoai le loc trung: nhung nhom da duoc nguoi dung xac nhan la KHONG trung
# ------------------------------------------------------------------

def add_dedup_exception(key_type, key_value, ten_dai_dien=""):
    conn = get_conn()
    conn.execute(
        "INSERT OR REPLACE INTO dedup_exceptions (key_type, key_value, ten_dai_dien, created_at) "
        "VALUES (?, ?, ?, ?)",
        (key_type, key_value, ten_dai_dien, datetime.datetime.now().isoformat(timespec="seconds")))
    conn.commit()
    conn.close()


def remove_dedup_exception(key_type, key_value):
    conn = get_conn()
    conn.execute(
        "DELETE FROM dedup_exceptions WHERE key_type = ? AND key_value = ?",
        (key_type, key_value))
    conn.commit()
    conn.close()


def list_dedup_exceptions(key_type):
    conn = get_conn()
    rows = conn.execute(
        "SELECT key_type, key_value, ten_dai_dien, created_at FROM dedup_exceptions "
        "WHERE key_type = ? ORDER BY created_at DESC", (key_type,)).fetchall()
    conn.close()
    return rows


def record_count():
    conn = get_conn()
    n = conn.execute("SELECT COUNT(*) FROM patients").fetchone()[0]
    conn.close()
    return n


# ------------------------------------------------------------------
# Doc & chuan hoa Excel
# ------------------------------------------------------------------

def normalize_date_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        if value.hour == 0 and value.minute == 0:
            return value.strftime("%d/%m/%Y")
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, datetime.date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        # File .xls (xlrd) tra ve so nguyen dang float (vd nam sinh 1989.0)
        # - bo ".0" cho gon, giong hanh vi openpyxl (tra ve int that su).
        return str(int(value))
    return str(value).strip()


def extract_birth_year(raw):
    if not raw:
        return None
    this_year = datetime.datetime.now().year
    for y in re.findall(r"\d{4}", raw):
        yi = int(y)
        if 1900 <= yi <= this_year:
            return yi
    return None


def parse_kham_date_iso(raw):
    """Tim mau dd/mm/yyyy o bat ky vi tri nao trong chuoi - file nguon co ca
    dinh dang 'dd/mm/yyyy HH:MM' lan 'HH:MM dd/mm/yyyy'."""
    if not raw:
        return None
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", raw)
    if not m:
        return None
    d, mo, y = m.groups()
    try:
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    except ValueError:
        return None


def at(row, col):
    """row: tuple gia tri 1 dong (0-based); col: chi so cot 1-based hoac None."""
    if col is None or col < 1 or col > len(row):
        return None
    return row[col - 1]


# ------------------------------------------------------------------
# Nhap Excel co anh xa cot (mapping): file bao cao BKLN thuc te khong dung
# chung 1 mau co dinh - ten cot khac nhau ("Họ và tên" vs "Tên bệnh nhân"),
# thu tu cot khac nhau, co file tach Gioi tinh/Nam sinh thanh 2 cot con
# "Nam"/"Nữ" thay vi 1 cot text, co file nhieu sheet voi bo cuc khac nhau
# trong cung workbook. Thay vi doan mot mau co dinh (rui ro doc sai neu bo
# cuc khac), ung dung tu doan truoc (theo ten cot, khong theo vi tri) roi
# nguoi dung xac nhan/sua lai qua hop thoai truoc khi thuc su nhap - xem
# ImportMappingDialog trong app.py.
# ------------------------------------------------------------------

MAPPING_FIELDS = [
    ("tt", "STT", False),
    ("ho_ten", "Họ và tên", True),
    ("gioi_tinh", "Giới tính", False),
    ("nam_sinh_raw", "Năm sinh", False),
    ("ma_bhyt", "Số BHYT", False),
    ("so_cccd", "Số CCCD", False),
    ("dia_chi", "Địa chỉ", False),
    ("phuong_xa", "Phường/Xã", False),
    ("tinh_tp", "Tỉnh/TP", False),
    ("ngay_kham_raw", "Ngày khám", False),
    ("chan_doan", "Chẩn đoán", False),
    ("benh_kem_theo", "Bệnh kèm theo", False),
]

# Cac bien the ten cot thuong gap trong thuc te - dung de tu doan anh xa.
# Khop kieu "chua chuoi con" (vd "CHẨN ĐOÁN\n(ICD)" van khop alias "chẩn
# đoán"), uu tien alias dai/cu the hon khi 1 cot co the khop nhieu truong.
FIELD_ALIASES = {
    "tt": ["stt", "số tt", "số thứ tự"],
    "ho_ten": ["họ và tên", "họ tên", "tên bệnh nhân"],
    "gioi_tinh": ["giới tính"],
    "nam_sinh_raw": ["năm sinh", "ngày sinh", "tuổi"],
    "ma_bhyt": ["mã thẻ bhyt", "mã bhyt", "số bhyt", "bhyt"],
    "so_cccd": ["cccd/ hộ chiếu", "cccd/hộ chiếu", "căn cước công dân",
                "số căn cước", "số cccd", "cccd", "cmnd"],
    "dia_chi": ["địa chỉ hành chính", "địa chỉ bảo hiểm", "địa chỉ"],
    "phuong_xa": ["phường/xã", "phường xã", "xã/phường", "xã (p)", "phường (xã)"],
    "tinh_tp": ["tỉnh/thành phố", "tỉnh thành phố", "tỉnh/tp", "tỉnh tp", "tỉnh (tp)"],
    "ngay_kham_raw": ["ngày khám bệnh", "ngày khám", "thời gian tiếp nhận"],
    "benh_kem_theo": ["tên bệnh kèm theo", "bệnh kèm theo", "bệnh kèm thèm", "tên bệnh phụ"],
    "chan_doan": ["mã bệnh chính", "chẩn đoán", "tên bệnh"],
}


def _norm_header(text):
    return re.sub(r"\s+", " ", str(text or "").strip().lower())


class _OpenpyxlReader:
    """Doc file .xlsx/.xlsm. read_only=True de doc nhanh file lon (~16.000
    dong), chi truy cap tuan tu bang iter_rows."""

    def __init__(self, path):
        if openpyxl is None:
            raise RuntimeError("Chưa cài thư viện openpyxl. Chạy: pip install openpyxl")
        self._wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    @property
    def sheet_names(self):
        return list(self._wb.sheetnames)

    def iter_rows(self, sheet_name, min_row, max_row=None):
        ws = self._wb[sheet_name]
        yield from ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True)

    def close(self):
        self._wb.close()


class _XlrdReader:
    """Doc file .xls cu (dinh dang nhi phan Excel 97-2003) qua xlrd - openpyxl
    khong doc duoc dinh dang nay. Tu chuyen cac o kieu ngay/gio (XL_CELL_DATE)
    sang datetime.datetime giong het openpyxl, de phan con lai cua core.py
    (normalize_date_value...) khong can biet file goc la dinh dang nao."""

    def __init__(self, path):
        if xlrd is None:
            raise RuntimeError(
                "Chưa cài thư viện xlrd (cần để đọc file .xls cũ định dạng Excel "
                "97-2003). Chạy: pip install xlrd — hoặc mở file bằng Excel rồi "
                "lưu lại dưới dạng .xlsx trước khi nhập.")
        self._wb = xlrd.open_workbook(path)

    @property
    def sheet_names(self):
        return list(self._wb.sheet_names())

    def iter_rows(self, sheet_name, min_row, max_row=None):
        ws = self._wb.sheet_by_name(sheet_name)
        end = ws.nrows if max_row is None else min(max_row, ws.nrows)
        datemode = self._wb.datemode
        for r in range(min_row - 1, end):
            values = []
            for cell in ws.row(r):
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        values.append(xlrd.xldate.xldate_as_datetime(cell.value, datemode))
                    except (xlrd.xldate.XLDateError, ValueError):
                        values.append(cell.value)
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    values.append(None)
                else:
                    values.append(cell.value)
            yield tuple(values)

    def close(self):
        pass


def _open_reader(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        return _XlrdReader(path)
    return _OpenpyxlReader(path)


def list_sheet_names(path):
    reader = _open_reader(path)
    try:
        return reader.sheet_names
    finally:
        reader.close()


def _match_columns_to_fields(columns):
    """columns: {chi_so_cot: text}. Tra ve {field: chi_so_cot} - chi nhung
    field khop duoc; neu 2 field cung khop 1 cot (vd "Tên bệnh nhân" chua ca
    alias cua ho_ten lan alias "tên bệnh" cua chan_doan) thi alias dai/cu
    the hon thang, khong phu thuoc thu tu cot trong file."""
    candidates = []
    for col, text in columns.items():
        norm = _norm_header(text)
        if not norm:
            continue
        for field, aliases in FIELD_ALIASES.items():
            for alias in aliases:
                if alias in norm:
                    candidates.append((len(alias), field, col))
    candidates.sort(key=lambda t: -t[0])
    result = {}
    used_cols = set()
    for _, field, col in candidates:
        if field in result or col in used_cols:
            continue
        result[field] = col
        used_cols.add(col)
    return result


def detect_header_row(reader, sheet_name, max_scan_rows=15):
    """Do tim dong tieu de: chon dong co nhieu o khop voi ten cac truong da
    biet nhat, trong so cac dong dau. Neu khong dong nao khop duoc gi thi
    mac dinh dong 1 - nguoi dung tu sua lai trong hop thoai anh xa cot."""
    best_row, best_score = 1, -1
    for r, row in enumerate(reader.iter_rows(sheet_name, min_row=1, max_row=max_scan_rows), start=1):
        columns = {c: v for c, v in enumerate(row, start=1) if v and str(v).strip()}
        score = len(_match_columns_to_fields(columns))
        if score > best_score:
            best_row, best_score = r, score
    return best_row


def auto_map_columns(columns):
    matched = _match_columns_to_fields(columns)
    return {field: matched.get(field) for field, _, _ in MAPPING_FIELDS}


def detect_gender_split(reader, sheet_name, header_row, mapping):
    """Mot so mau bao cao tach Gioi tinh + Nam sinh/Tuoi thanh 2 cot con
    "Nam"/"Nữ" ngay duoi dong tieu de (gia tri nam o cot nao xac dinh gioi
    tinh, xem README). Chi goi ham nay khi CHUA tu khop duoc cot "Giới
    tính" rieng - tra ve (cot_nam, cot_nu) hoac (None, None)."""
    if mapping.get("gioi_tinh"):
        return None, None
    sub_row = next(reader.iter_rows(sheet_name, min_row=header_row + 1, max_row=header_row + 1), None)
    if not sub_row:
        return None, None
    male_col = female_col = None
    for c, val in enumerate(sub_row, start=1):
        norm = _norm_header(val)
        if norm == "nam":
            male_col = c
        elif norm in ("nữ", "nu"):
            female_col = c
    if male_col and female_col:
        return male_col, female_col
    return None, None


def detect_sheet_mapping(path, sheet_name, header_row=None):
    """Doan toan bo thong tin can de nhap 1 sheet: dong tieu de, danh sach
    cot, anh xa tu dong theo ten cot, co tach Gioi tinh theo Nam/Nữ khong,
    va vai dong du lieu dau de xem truoc. Goi lai voi header_row cu the khi
    nguoi dung tu sua lai dong tieu de trong hop thoai."""
    reader = _open_reader(path)
    try:
        if header_row is None:
            header_row = detect_header_row(reader, sheet_name)
        header_cells = next(reader.iter_rows(sheet_name, min_row=header_row, max_row=header_row), ())
        columns = {c: str(v).strip() for c, v in enumerate(header_cells, start=1) if v and str(v).strip()}
        mapping = auto_map_columns(columns)
        male_col, female_col = detect_gender_split(reader, sheet_name, header_row, mapping)
        gender_split_cols = (male_col, female_col) if male_col and female_col else None
        preview_start = header_row + (2 if gender_split_cols else 1)
        preview_rows = []
        for i, row in enumerate(reader.iter_rows(sheet_name, min_row=preview_start)):
            if i >= 5:
                break
            preview_rows.append(row)
    finally:
        reader.close()
    return {
        "header_row": header_row,
        "columns": columns,
        "mapping": mapping,
        "gender_split_cols": gender_split_cols,
        "preview_rows": preview_rows,
    }


def read_excel_rows_mapped(path, sheet_name, header_row, mapping, gender_split_cols=None):
    """Sinh ra tung dict du lieu tho cho moi dong benh nhan, theo dung anh
    xa cot (mapping) da duoc xac nhan qua hop thoai - khong con doan vi tri
    cot theo offset co dinh nhu truoc."""
    reader = _open_reader(path)

    col_tt = mapping.get("tt")
    col_hoten = mapping.get("ho_ten")
    col_gioitinh = mapping.get("gioi_tinh")
    col_ngaysinh = mapping.get("nam_sinh_raw")
    col_bhyt = mapping.get("ma_bhyt")
    col_cccd = mapping.get("so_cccd")
    col_diachi = mapping.get("dia_chi")
    col_phuongxa = mapping.get("phuong_xa")
    col_tinhtp = mapping.get("tinh_tp")
    col_ngaykham = mapping.get("ngay_kham_raw")
    col_chandoan = mapping.get("chan_doan")
    col_benhkem = mapping.get("benh_kem_theo")
    male_col, female_col = gender_split_cols or (None, None)

    # Neu co dong phu Nam/Nữ ngay duoi dong tieu de (xem detect_gender_split),
    # bo qua dong do khi doc du lieu that su.
    data_start = header_row + (2 if (male_col and female_col) else 1)
    auto_tt = 0
    try:
        for row in reader.iter_rows(sheet_name, min_row=data_start):
            ho_ten = at(row, col_hoten)
            so_cccd = at(row, col_cccd)
            if not ho_ten and not so_cccd:
                continue
            auto_tt += 1

            if male_col and female_col:
                male_val = at(row, male_col)
                female_val = at(row, female_col)
                if male_val not in (None, ""):
                    gioi_tinh, nam_sinh_raw = "Nam", normalize_date_value(male_val)
                elif female_val not in (None, ""):
                    gioi_tinh, nam_sinh_raw = "Nữ", normalize_date_value(female_val)
                else:
                    gioi_tinh, nam_sinh_raw = "", ""
            else:
                gioi_tinh = str(at(row, col_gioitinh) or "").strip()
                nam_sinh_raw = normalize_date_value(at(row, col_ngaysinh))
                # Mot so dong trong file nguon bi dao nham cot Gioi tinh <-> Nam
                # sinh (vi du gioi_tinh='1958', nam_sinh_raw='Nam'). Phat hien
                # va hoan lai.
                if gioi_tinh not in ("Nam", "Nữ") and nam_sinh_raw in ("Nam", "Nữ"):
                    gioi_tinh, nam_sinh_raw = nam_sinh_raw, gioi_tinh

            ngay_kham_raw = normalize_date_value(at(row, col_ngaykham))
            yield {
                "tt": at(row, col_tt) if col_tt else auto_tt,
                "ho_ten": (str(ho_ten).strip() if ho_ten else ""),
                "gioi_tinh": gioi_tinh,
                "nam_sinh_raw": nam_sinh_raw,
                "birth_year": extract_birth_year(nam_sinh_raw),
                "ma_bhyt": (str(at(row, col_bhyt) or "").strip()),
                "so_cccd": (str(so_cccd or "").strip()),
                "dia_chi": (str(at(row, col_diachi) or "").strip()),
                "phuong_xa": (str(at(row, col_phuongxa) or "").strip()),
                "tinh_tp": (str(at(row, col_tinhtp) or "").strip()),
                "ngay_kham_raw": ngay_kham_raw,
                "ngay_kham_date": parse_kham_date_iso(ngay_kham_raw),
                "chan_doan": (str(at(row, col_chandoan) or "").strip()),
                "benh_kem_theo": (str(at(row, col_benhkem) or "").strip()),
            }
    finally:
        reader.close()


def row_hash(r):
    key = "|".join([
        (r.get("ho_ten") or "").lower(),
        (r.get("gioi_tinh") or "").lower(),
        (r.get("nam_sinh_raw") or "").lower(),
        (r.get("ma_bhyt") or "").lower(),
        (r.get("so_cccd") or "").lower(),
        (r.get("dia_chi") or "").lower(),
        (r.get("phuong_xa") or "").lower(),
        (r.get("tinh_tp") or "").lower(),
        (r.get("ngay_kham_raw") or "").lower(),
        (r.get("chan_doan") or "").lower(),
        (r.get("benh_kem_theo") or "").lower(),
    ])
    return hashlib.sha256(key.encode("utf-8")).hexdigest()


def import_excel_mapped(path, sheet_name, header_row, mapping, gender_split_cols=None, progress_cb=None):
    """Doc 1 sheet cua file Excel theo dung anh xa cot (mapping) da xac
    nhan va nhap vao SQLite. Tra ve (tong_doc, them_moi, bo_qua_trung). Ghi
    theo lo (executemany moi 500 dong) de giam so lan goi mang khi dang o
    che do may tram (moi lan goi la 1 request qua mang toi may chu)."""
    conn = get_conn()
    cur = conn.cursor()
    total = 0
    inserted = 0
    nguon = os.path.basename(path)
    now = datetime.datetime.now().isoformat(timespec="seconds")
    sql = """
        INSERT OR IGNORE INTO patients
        (tt, ho_ten, gioi_tinh, nam_sinh_raw, birth_year, ma_bhyt, so_cccd,
         dia_chi, phuong_xa, tinh_tp, ngay_kham_raw, ngay_kham_date,
         chan_doan, benh, benh_kem_theo, nguon_file, imported_at, row_hash)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """
    batch = []

    def flush():
        nonlocal inserted
        if not batch:
            return
        cur.executemany(sql, batch)
        if cur.rowcount and cur.rowcount > 0:
            inserted += cur.rowcount
        batch.clear()

    for r in read_excel_rows_mapped(path, sheet_name, header_row, mapping, gender_split_cols):
        total += 1
        h = row_hash(r)
        batch.append((
            r["tt"], r["ho_ten"], r["gioi_tinh"], r["nam_sinh_raw"], r["birth_year"],
            r["ma_bhyt"], r["so_cccd"], r["dia_chi"], r["phuong_xa"], r["tinh_tp"],
            r["ngay_kham_raw"], r["ngay_kham_date"], r["chan_doan"],
            infer_diseases(r["chan_doan"]), r["benh_kem_theo"],
            nguon, now, h,
        ))
        if len(batch) >= 500:
            flush()
        if progress_cb and total % 500 == 0:
            progress_cb(total)
    flush()
    conn.commit()
    conn.close()
    return total, inserted, total - inserted


# ------------------------------------------------------------------
# Xuat CSV / Excel
# ------------------------------------------------------------------

def write_export(path, headers, rows):
    """Ghi headers + rows (moi row la sequence) ra file. Dinh dang xac dinh
    theo phan mo rong cua path: .xlsx dung openpyxl, con lai mac dinh CSV."""
    ext = os.path.splitext(path)[1].lower()
    n = 0
    if ext == ".xlsx":
        if openpyxl is None:
            raise RuntimeError("Chưa cài thư viện openpyxl để xuất Excel. Chạy: pip install openpyxl")
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("Danh sách")
        ws.append(list(headers))
        for row in rows:
            ws.append(list(row))
            n += 1
        wb.save(path)
    else:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for row in rows:
                w.writerow(list(row))
                n += 1
    return n


def export_query_to_file(sql, params, path, headers=None):
    conn = get_conn()
    cur = conn.execute(sql, params or [])
    cols = headers or [d[0] for d in cur.description]
    n = write_export(path, cols, cur)
    conn.close()
    return n


# ------------------------------------------------------------------
# Loc trung: quet, xem chi tiet nhom, xuat danh sach duy nhat
# (co the gop lich su kham thay vi chi giu 1 dong va bo cac dong con lai)
# ------------------------------------------------------------------

def scan_dedup_groups(key_expr, key_where, key_type):
    conn = get_conn()
    rows = conn.execute(f"""
        SELECT {key_expr} AS k, COUNT(*) AS n, MIN(ho_ten) AS ten
        FROM patients
        WHERE {key_where}
          AND {key_expr} NOT IN (
              SELECT key_value FROM dedup_exceptions WHERE key_type = ?
          )
        GROUP BY k
        HAVING COUNT(*) > 1
        ORDER BY n DESC
    """, (key_type,)).fetchall()
    conn.close()
    return rows


def group_detail_rows(key_expr, key_value):
    conn = get_conn()
    rows = conn.execute(
        f"SELECT {', '.join(c for c, _ in COLUMNS)} FROM patients "
        f"WHERE {key_expr} = ? ORDER BY id", (key_value,)
    ).fetchall()
    conn.close()
    return rows


def delete_patients_by_ids(ids):
    if not ids:
        return 0
    conn = get_conn()
    conn.executemany("DELETE FROM patients WHERE id = ?", [(i,) for i in ids])
    conn.commit()
    conn.close()
    return len(ids)


# ------------------------------------------------------------------
# Gop trung: giu lai 1 ban ghi "chinh", cac ban ghi con lai bi gop vao
# cot lich_su_kham cua ban ghi chinh (khong mat thong tin lich su kham benh)
# roi moi bi xoa - khac voi xoa thang (mat het thong tin cac lan kham cu).
# ------------------------------------------------------------------

def _history_entry(row):
    ngay = (row["ngay_kham_raw"] or "").strip() or "Không rõ ngày"
    chan_doan = (row["chan_doan"] or "").strip() or "Không rõ chẩn đoán"
    return f"{ngay}: {chan_doan}"


def _row_history_entries(row):
    """Neu ban ghi nay da tung duoc gop truoc do (da co lich_su_kham) thi lay
    lai cac muc da luu, khong thi coi ban than no la 1 muc lich su."""
    existing = (row["lich_su_kham"] or "").strip()
    if existing:
        return [e.strip() for e in existing.split("  |  ") if e.strip()]
    return [_history_entry(row)]


def merge_rows(primary_id, other_ids):
    """Gop other_ids vao primary_id: cong don lich su kham vao ban ghi chinh
    (primary), roi xoa cac ban ghi con lai. Tra ve so ban ghi da gop."""
    other_ids = [i for i in other_ids if i != primary_id]
    if not other_ids:
        return 0
    conn = get_conn()
    all_ids = [primary_id] + other_ids
    placeholders = ",".join("?" * len(all_ids))
    rows = conn.execute(
        f"SELECT id, ngay_kham_raw, chan_doan, lich_su_kham FROM patients "
        f"WHERE id IN ({placeholders})", all_ids
    ).fetchall()
    by_id = {r["id"]: r for r in rows}

    entries = []
    for i in all_ids:
        r = by_id.get(i)
        if r is not None:
            entries.extend(_row_history_entries(r))
    seen = set()
    unique_entries = []
    for e in entries:
        if e not in seen:
            seen.add(e)
            unique_entries.append(e)
    merged_history = "  |  ".join(unique_entries)

    conn.execute("UPDATE patients SET lich_su_kham = ? WHERE id = ?", (merged_history, primary_id))
    conn.executemany("DELETE FROM patients WHERE id = ?", [(i,) for i in other_ids])
    conn.commit()
    conn.close()
    return len(other_ids)


def merge_specific_ids(ids, order_by):
    """Gop 1 danh sach id ban ghi tuy chon (nguoi dung tu tich chon) thanh
    1 ban ghi, chon ban ghi chinh theo order_by. Tra ve so ban ghi da gop."""
    ids = list(dict.fromkeys(ids))
    if len(ids) < 2:
        return 0
    conn = get_conn()
    placeholders = ",".join("?" * len(ids))
    rows = conn.execute(
        f"SELECT id FROM patients WHERE id IN ({placeholders}) {order_by}", ids
    ).fetchall()
    conn.close()
    ordered_ids = [r["id"] for r in rows]
    if len(ordered_ids) < 2:
        return 0
    return merge_rows(ordered_ids[0], ordered_ids[1:])


def merge_group(key_expr, key_value, order_by):
    """Gop toan bo 1 nhom trung thanh 1 ban ghi (ban ghi chinh chon theo
    order_by). Tra ve so ban ghi da gop (khong tinh ban ghi chinh)."""
    conn = get_conn()
    rows = conn.execute(
        f"SELECT id FROM patients WHERE {key_expr} = ? {order_by}", (key_value,)
    ).fetchall()
    conn.close()
    ids = [r["id"] for r in rows]
    if len(ids) < 2:
        return 0
    return merge_rows(ids[0], ids[1:])


def merge_all_groups(key_expr, key_where, key_type, order_by):
    """Gop tat ca cac nhom trung hien dang duoc quet ra (da tru cac nhom
    duoc xac nhan khong trung), bang SQL theo lo (khong lap tung nhom trong
    Python) de xu ly nhanh ngay ca khi co hang nghin nhom. Tra ve
    (so_nhom_da_gop, so_ban_ghi_da_gop)."""
    conn = get_conn()
    try:
        conn.execute("DROP TABLE IF EXISTS temp._merge_ranked")
        conn.execute("DROP TABLE IF EXISTS temp._merge_groups")
        conn.execute("DROP TABLE IF EXISTS temp._merge_history")

        conn.execute(f"""
            CREATE TEMP TABLE _merge_ranked AS
            SELECT id, {key_expr} AS k,
                   ROW_NUMBER() OVER (PARTITION BY {key_expr} {order_by}) AS rn
            FROM patients
            WHERE {key_where}
              AND {key_expr} NOT IN (SELECT key_value FROM dedup_exceptions WHERE key_type = ?)
        """, (key_type,))
        conn.execute("CREATE INDEX _idx_mr_id ON _merge_ranked(id)")
        conn.execute("CREATE INDEX _idx_mr_k ON _merge_ranked(k)")

        conn.execute("""
            CREATE TEMP TABLE _merge_groups AS
            SELECT k FROM _merge_ranked GROUP BY k HAVING COUNT(*) > 1
        """)
        n_groups = conn.execute("SELECT COUNT(*) FROM _merge_groups").fetchone()[0]
        if n_groups == 0:
            return 0, 0

        conn.execute("""
            CREATE TEMP TABLE _merge_history AS
            SELECT k, GROUP_CONCAT(entry, '  |  ') AS lich_su FROM (
                SELECT mr.k AS k,
                       CASE WHEN TRIM(IFNULL(p.lich_su_kham, '')) <> '' THEN p.lich_su_kham
                            ELSE (COALESCE(NULLIF(TRIM(p.ngay_kham_raw), ''), 'Không rõ ngày') || ': ' ||
                                  COALESCE(NULLIF(TRIM(p.chan_doan), ''), 'Không rõ chẩn đoán'))
                       END AS entry
                FROM _merge_ranked mr
                JOIN patients p ON p.id = mr.id
                WHERE mr.k IN (SELECT k FROM _merge_groups)
                ORDER BY mr.k, (p.ngay_kham_date IS NULL), p.ngay_kham_date DESC, p.id DESC
            )
            GROUP BY k
        """)

        conn.execute("""
            UPDATE patients
            SET lich_su_kham = (SELECT lich_su FROM _merge_history mh WHERE mh.k = mr.k)
            FROM _merge_ranked mr
            WHERE patients.id = mr.id AND mr.rn = 1 AND mr.k IN (SELECT k FROM _merge_groups)
        """)

        cur = conn.execute("""
            DELETE FROM patients WHERE id IN (
                SELECT id FROM _merge_ranked WHERE rn > 1 AND k IN (SELECT k FROM _merge_groups)
            )
        """)
        n_merged = cur.rowcount
        conn.commit()
        return n_groups, n_merged
    finally:
        conn.execute("DROP TABLE IF EXISTS temp._merge_ranked")
        conn.execute("DROP TABLE IF EXISTS temp._merge_groups")
        conn.execute("DROP TABLE IF EXISTS temp._merge_history")
        conn.close()


def unique_rows_with_optional_history(key_expr, key_where, order_by, include_history):
    """Tra ve (headers, rows) danh sach da loc trung - 1 dong/nguoi.
    Neu include_history=True, them cot 'Lịch sử khám' gop tat ca cac lan kham
    (ngay kham: chan doan) cua nguoi do, moi nhat truoc, thay vi bo cac dong con lai."""
    conn = get_conn()
    col_list = ", ".join(c for c, _ in COLUMNS)
    ranked_cols = ", ".join(f"ranked.{c}" for c, _ in COLUMNS)
    headers = [label for _, label in COLUMNS]

    if include_history:
        rows = conn.execute(f"""
            WITH ranked AS (
                SELECT {col_list},
                       ROW_NUMBER() OVER (PARTITION BY {key_expr} {order_by}) AS rn,
                       {key_expr} AS kexpr
                FROM patients
            ),
            history AS (
                SELECT kk, GROUP_CONCAT(entry, '  |  ') AS lich_su FROM (
                    SELECT {key_expr} AS kk,
                           (COALESCE(NULLIF(TRIM(ngay_kham_raw), ''), 'Không rõ ngày') || ': ' ||
                            COALESCE(NULLIF(TRIM(chan_doan), ''), 'Không rõ chẩn đoán')) AS entry
                    FROM patients
                    WHERE {key_where}
                    ORDER BY {key_expr}, (ngay_kham_date IS NULL), ngay_kham_date DESC, id DESC
                )
                GROUP BY kk
            )
            SELECT {ranked_cols}, history.lich_su
            FROM ranked LEFT JOIN history ON history.kk = ranked.kexpr
            WHERE ranked.kexpr IS NULL OR TRIM(ranked.kexpr) = '' OR ranked.rn = 1
            ORDER BY ranked.id
        """).fetchall()
        headers = headers + ["Lịch sử khám"]
    else:
        rows = conn.execute(f"""
            WITH ranked AS (
                SELECT {col_list},
                       ROW_NUMBER() OVER (PARTITION BY {key_expr} {order_by}) AS rn,
                       {key_expr} AS kexpr
                FROM patients
            )
            SELECT {col_list} FROM ranked
            WHERE kexpr IS NULL OR TRIM(kexpr) = '' OR rn = 1
            ORDER BY id
        """).fetchall()
    conn.close()
    return headers, rows


# ------------------------------------------------------------------
# Sua loi du lieu: cot Gioi tinh <-> Ngay sinh bi dao cho trong file nguon
# ------------------------------------------------------------------

def count_swapped_gender_birthdate():
    conn = get_conn()
    n = conn.execute(
        "SELECT COUNT(*) FROM patients "
        "WHERE gioi_tinh NOT IN ('Nam','Nữ') AND nam_sinh_raw IN ('Nam','Nữ')"
    ).fetchone()[0]
    conn.close()
    return n


def fix_swapped_gender_birthdate():
    """Sua cac dong ma cot Gioi tinh dang chua gia tri Ngay sinh va nguoc lai
    (loi nhap lieu tai file Excel nguon). Tra ve so dong da sua."""
    conn = get_conn()
    rows = conn.execute(
        "SELECT id, gioi_tinh, nam_sinh_raw FROM patients "
        "WHERE gioi_tinh NOT IN ('Nam','Nữ') AND nam_sinh_raw IN ('Nam','Nữ')"
    ).fetchall()
    for r in rows:
        new_gioi = r["nam_sinh_raw"]
        new_nam_sinh = r["gioi_tinh"]
        conn.execute(
            "UPDATE patients SET gioi_tinh=?, nam_sinh_raw=?, birth_year=? WHERE id=?",
            (new_gioi, new_nam_sinh, extract_birth_year(new_nam_sinh), r["id"])
        )
    conn.commit()
    conn.close()
    return len(rows)


def count_unparsed_kham_dates():
    conn = get_conn()
    n = conn.execute(
        "SELECT COUNT(*) FROM patients WHERE ngay_kham_date IS NULL "
        "AND ngay_kham_raw IS NOT NULL AND TRIM(ngay_kham_raw) <> ''"
    ).fetchone()[0]
    conn.close()
    return n


def fix_unparsed_kham_dates():
    """Tinh lai ngay_kham_date cho cac dong bi bo sot truoc khi parse_kham_date_iso
    duoc sua de nhan dang ca dinh dang 'HH:MM dd/mm/yyyy'. Tra ve so dong da sua."""
    conn = get_conn()
    rows = conn.execute(
        "SELECT id, ngay_kham_raw FROM patients WHERE ngay_kham_date IS NULL "
        "AND ngay_kham_raw IS NOT NULL AND TRIM(ngay_kham_raw) <> ''"
    ).fetchall()
    fixed = 0
    for r in rows:
        new_date = parse_kham_date_iso(r["ngay_kham_raw"])
        if new_date:
            conn.execute("UPDATE patients SET ngay_kham_date = ? WHERE id = ?", (new_date, r["id"]))
            fixed += 1
    conn.commit()
    conn.close()
    return fixed


def count_unclassified_diseases():
    """So dong co Chan doan nhung chua duoc gan Nhom benh (cot 'benh' rong) -
    thuong la du lieu nhap tu truoc khi co tinh nang nay."""
    conn = get_conn()
    n = conn.execute(
        "SELECT COUNT(*) FROM patients WHERE TRIM(IFNULL(chan_doan,'')) <> '' "
        "AND TRIM(IFNULL(benh,'')) = ''"
    ).fetchone()[0]
    conn.close()
    return n


def reclassify_diseases(overwrite=False):
    """Tinh lai cot 'benh' tu 'chan_doan' bang tu khoa (xem DISEASE_CATEGORIES).
    Mac dinh (overwrite=False) chi tinh cho cac dong dang co Nhom benh rong,
    de khong ghi de nhung dong nguoi dung da tu xem lai/sua tay. Tra ve so
    dong da cap nhat."""
    conn = get_conn()
    if overwrite:
        rows = conn.execute(
            "SELECT id, chan_doan FROM patients WHERE TRIM(IFNULL(chan_doan,'')) <> ''"
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, chan_doan FROM patients WHERE TRIM(IFNULL(chan_doan,'')) <> '' "
            "AND TRIM(IFNULL(benh,'')) = ''"
        ).fetchall()
    updated = 0
    for r in rows:
        new_benh = infer_diseases(r["chan_doan"])
        if new_benh:
            conn.execute("UPDATE patients SET benh = ? WHERE id = ?", (new_benh, r["id"]))
            updated += 1
    conn.commit()
    conn.close()
    return updated


# ------------------------------------------------------------------
# Sao luu tu dong truoc cac thao tac nguy hiem (gop hang loat, xoa, reset)
# ------------------------------------------------------------------

def backup_database(reason="thao_tac", keep=10):
    """Sao chep benh_nhan.db sang thu muc backups/ kem timestamp. Tu dong don
    bot, chi giu lai `keep` ban gan nhat. Tra ve duong dan file backup, hoac
    None neu chua co CSDL de sao luu.
    O che do may tram, chuyen tiep yeu cau nay sang may chu (chi may chu moi
    thuc su giu file benh_nhan.db va thu muc backups/)."""
    if REMOTE_BASE_URL:
        from netclient import request_backup
        return request_backup(REMOTE_BASE_URL, reason, keep, REMOTE_API_KEY)
    if not os.path.exists(DB_PATH):
        return None
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_reason = re.sub(r"[^0-9a-zA-Z_]+", "_", reason)
    dest = os.path.join(BACKUP_DIR, f"benh_nhan_{ts}_{safe_reason}.db")
    shutil.copy2(DB_PATH, dest)

    backups = sorted(
        f for f in os.listdir(BACKUP_DIR)
        if f.startswith("benh_nhan_") and f.endswith(".db")
    )
    for old in backups[:-keep]:
        try:
            os.remove(os.path.join(BACKUP_DIR, old))
        except OSError:
            pass
    return dest


def list_backups():
    if not os.path.isdir(BACKUP_DIR):
        return []
    return sorted(
        f for f in os.listdir(BACKUP_DIR)
        if f.startswith("benh_nhan_") and f.endswith(".db")
    )


# ------------------------------------------------------------------
# Bao cao chat luong du lieu
# ------------------------------------------------------------------

DATA_QUALITY_CHECKS = [
    ("Thiếu Số CCCD", "(so_cccd IS NULL OR TRIM(so_cccd) = '')"),
    ("Thiếu Mã BHYT", "(ma_bhyt IS NULL OR TRIM(ma_bhyt) = '')"),
    ("Thiếu Chẩn đoán", "(chan_doan IS NULL OR TRIM(chan_doan) = '')"),
    ("Không xác định được Năm sinh", "(birth_year IS NULL)"),
    ("Không xác định được Ngày khám", "(ngay_kham_date IS NULL)"),
]


def data_quality_summary():
    """Tra ve list (nhan, so_dong) cho tung loai van de du lieu."""
    conn = get_conn()
    result = []
    for label, cond in DATA_QUALITY_CHECKS:
        n = conn.execute(f"SELECT COUNT(*) FROM patients WHERE {cond}").fetchone()[0]
        result.append((label, n))
    conn.close()
    return result


def data_quality_rows_sql():
    """Cau SQL tra ve cac dong co van de, kem cot 'loai_loi' liet ke ly do."""
    case_parts = " || ".join(
        f"CASE WHEN {cond} THEN '{label}; ' ELSE '' END" for label, cond in DATA_QUALITY_CHECKS
    )
    where_parts = " OR ".join(cond for _, cond in DATA_QUALITY_CHECKS)
    cols = ", ".join(c for c, _ in COLUMNS)
    return f"""
        SELECT {cols}, TRIM({case_parts}) AS loai_loi
        FROM patients
        WHERE {where_parts}
        ORDER BY id
    """


# ------------------------------------------------------------------
# Thong ke truc quan (dung cho tab "Thong ke" - ve bieu do bang QtCharts)
# ------------------------------------------------------------------

# Chi cho phep thong ke theo cac cot trong danh sach nay (dua thang ten cot
# vao f-string SQL ben duoi - KHONG duoc nhan gia tri tuy y tu nguoi dung,
# chi duoc goi voi 1 trong cac key co san o day).
STATS_COLUMNS = {
    "gioi_tinh": "Giới tính",
    "tinh_tp": "Tỉnh/Thành phố",
    "phuong_xa": "Phường/Xã",
    "chan_doan": "Chẩn đoán",
}


def stats_top_values(column, limit=20):
    """Tra ve list (gia_tri, so_luong) cho 1 cot trong STATS_COLUMNS, xep
    theo so luong giam dan (bo qua gia tri rong). Dung de ve bieu do."""
    if column not in STATS_COLUMNS:
        raise ValueError(f"Cột không hợp lệ để thống kê: {column}")
    conn = get_conn()
    rows = conn.execute(f"""
        SELECT {column}, COUNT(*) AS n FROM patients
        WHERE TRIM(IFNULL({column}, '')) <> ''
        GROUP BY {column} ORDER BY n DESC, {column} LIMIT ?
    """, (limit,)).fetchall()
    conn.close()
    return [(r[0], r[1]) for r in rows]


def stats_birth_decade():
    """Tra ve list (nhan_thap_ky vd '1960s', so_luong) theo thap ky nam sinh,
    sap xep tang dan. Dung de ve bieu do phan bo do tuoi."""
    conn = get_conn()
    rows = conn.execute("""
        SELECT (birth_year / 10) * 10 AS decade, COUNT(*) AS n
        FROM patients WHERE birth_year IS NOT NULL
        GROUP BY decade ORDER BY decade
    """).fetchall()
    conn.close()
    return [(f"{int(r[0])}s", r[1]) for r in rows]


def stats_disease_counts():
    """Tra ve list (nhan_nhom_benh, so_luong) - dem theo TUNG nhom benh
    rieng le, khong phai theo to hop. Vi 1 benh nhan co the thuoc nhieu
    nhom cung luc (cot 'benh' luu dang 'THA, ĐTĐ'), tach tung nhan trong
    Python thay vi GROUP BY truc tiep tren cot (se gop nham cac to hop
    khac nhau thanh nhom rieng). Sap xep theo so luong giam dan."""
    conn = get_conn()
    rows = conn.execute("SELECT benh FROM patients WHERE TRIM(IFNULL(benh, '')) <> ''").fetchall()
    conn.close()
    counts = {}
    for r in rows:
        for label in (r["benh"] or "").split(","):
            label = label.strip()
            if label:
                counts[label] = counts.get(label, 0) + 1
    return sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))


# ------------------------------------------------------------------
# Mat khau bao ve ung dung (kiem soat truy cap giao dien - KHONG ma hoa
# file benh_nhan.db; ai co file van mo duoc bang cong cu SQLite khac)
# ------------------------------------------------------------------

def has_password():
    return os.path.exists(PASSWORD_FILE)


def _hash_password(password, salt):
    return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200_000).hex()


def set_password(password):
    salt = secrets.token_bytes(16)
    h = _hash_password(password, salt)
    with open(PASSWORD_FILE, "w", encoding="utf-8") as f:
        f.write(salt.hex() + ":" + h)


def verify_password(password):
    if not has_password():
        return True
    with open(PASSWORD_FILE, "r", encoding="utf-8") as f:
        salt_hex, h = f.read().strip().split(":", 1)
    return _hash_password(password, bytes.fromhex(salt_hex)) == h


def remove_password():
    if os.path.exists(PASSWORD_FILE):
        os.remove(PASSWORD_FILE)


# ------------------------------------------------------------------
# Kiem tra ban cap nhat (khong chan giao dien - dung khi khoi dong app)
# ------------------------------------------------------------------

def get_local_version():
    if os.path.exists(VERSION_FILE):
        with open(VERSION_FILE, "r", encoding="utf-8") as f:
            return f.read().strip()
    return None


def get_update_token():
    if os.path.exists(UPDATE_TOKEN_FILE):
        with open(UPDATE_TOKEN_FILE, "r", encoding="utf-8") as f:
            return f.read().strip()
    return None


def check_latest_release(timeout=5, tag_prefix="v"):
    """Tra ve (tag_moi_nhat, html_url) cua release moi nhat co tag bat dau
    bang `tag_prefix`, hoac (None, None) neu khong kiem tra duoc (khong
    mang, repo chua co release phu hop...). Khong bao gio raise loi - danh
    cho kiem tra nen, khong duoc lam gian doan app.

    Chi con 1 dong phien ban duy nhat (tag "vX.Y.Z", dung chung cho ca ung
    dung chinh lan thanh phan May chu - xem setup.iss). Dung /releases
    (danh sach) thay vi /releases/latest de loc dung tien to tag can, vi
    /releases/latest tra ve release moi nhat theo thoi gian bat ke tag nao.

    Token la tuy chon - chi can neu repo dang o che do Private (xem
    get_update_token()); repo Public van kiem tra duoc binh thuong ma
    khong can token."""
    token = get_update_token()
    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/releases?per_page=20"
    headers = {
        "Accept": "application/vnd.github+json",
        "User-Agent": "QuanLyBenhNhanTHA",
    }
    if token:
        headers["Authorization"] = f"token {token}"
    req = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            releases = json.loads(resp.read().decode("utf-8"))
        for r in releases:
            tag = r.get("tag_name") or ""
            if tag.startswith(tag_prefix) and not r.get("draft") and not r.get("prerelease"):
                return tag, r.get("html_url")
        return None, None
    except Exception:
        return None, None
